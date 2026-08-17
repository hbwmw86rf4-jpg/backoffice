import openpyxl
import sqlite3
import os
import sys
import xml.etree.ElementTree as ET

sys.stdout.reconfigure(encoding='utf-8')

def clean_tag(elem):
    return elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag

def get_official_fgm_fuel(staging_dir, target_date):
    if not os.path.exists(staging_dir):
        return None
    fgm_files = [f for f in os.listdir(staging_dir) if f.upper().startswith('FGM') and f.upper().endswith('.XML')]
    result = {'001': {'amt': 0.0, 'gal': 0.0, 'disc': 0.0}, '002': {'amt': 0.0, 'gal': 0.0, 'disc': 0.0}, '003': {'amt': 0.0, 'gal': 0.0, 'disc': 0.0}}
    found = False

    for fn in fgm_files:
        fpath = os.path.join(staging_dir, fn)
        try:
            tree = ET.parse(fpath)
            root = tree.getroot()
            bdate = ''
            for elem in root.iter():
                if clean_tag(elem) in ['JournalHeader', 'MovementHeader']:
                    for child in elem.iter():
                        if clean_tag(child) == 'BusinessDate' and child.text:
                            bdate = child.text
            if bdate != target_date: continue

            for elem in root.iter():
                if clean_tag(elem) == 'FGMDetail':
                    grade_id = ''
                    for child in elem:
                        if clean_tag(child) == 'FuelGradeID' and child.text:
                            grade_id = child.text
                        elif clean_tag(child) == 'FGMSalesTotals':
                            vol = 0.0
                            amt = 0.0
                            disc = 0.0
                            for st_child in child:
                                stag = clean_tag(st_child)
                                if stag == 'FuelGradeSalesVolume' and st_child.text: vol = float(st_child.text)
                                elif stag == 'FuelGradeSalesAmount' and st_child.text: amt = float(st_child.text)
                                elif stag == 'DispenserDiscountAmount' and st_child.text: disc = float(st_child.text)

                            if grade_id in result and (vol > 0 or amt > 0):
                                found = True
                                result[grade_id]['gal'] += vol
                                result[grade_id]['amt'] += amt
                                result[grade_id]['disc'] += disc
        except Exception as e:
            pass

    return result if found else None

def generate_recon(target_date=None, out_dir=None):
    base_dir = r'C:\Users\shell\Documents\office\backoffice'
    db_path = os.path.join(base_dir, 'data', 'backoffice.db')
    staging_dir = os.path.join(base_dir, 'data', 'staging', 'BOOutBox')
    template_path = os.path.join(base_dir, 'data', 'templates', 'RochesterShell_DailyRecon.xlsx')
    
    if not out_dir:
        out_dir = os.path.join(base_dir, 'data', 'recon')
    os.makedirs(out_dir, exist_ok=True)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    if not target_date:
        cur.execute("SELECT MAX(business_date) FROM transactions")
        row = cur.fetchone()
        target_date = row[0] if row and row[0] else '2026-08-15'

    output_path = os.path.join(out_dir, f'RochesterShell_DailyRecon_{target_date}.xlsx')

    if not os.path.exists(template_path):
        print(f"Template not found: {template_path}")
        return None

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Daily Reconciliation']

    # Date
    ws['I2'] = target_date

    # Department mapping
    dept_mapping = {
        'Auto Parts': 5,
        'Beer': 6,
        'Candy': 7,
        'Cig Cartons': 8,
        'Cigs': 9,
        'Coffee': 10,
        'Deli': 11,
        'Edible': 12,
        'Fountain': 13,
        'GROC No Tax': 14,
        'Ice': 15,
        'Liquor': 16,
        'Non-Edible': 17,
        'Snacks': 18,
        'Soda': 19,
        'Tobacco': 20,
        'Vapes': 21,
        'HBA': 22
    }

    cur.execute("""
        SELECT COALESCE(d.name, 'Dept ' || ti.merchandise_code), SUM(ti.total_amount)
        FROM transaction_items ti
        JOIN transactions t ON ti.transaction_id = t.id
        LEFT JOIN pricebook pb ON ti.upc = pb.upc
        LEFT JOIN departments d ON d.id = COALESCE(pb.department_id, CAST(ti.merchandise_code AS INTEGER))
        WHERE t.business_date = ? AND ti.item_type = 'cstore'
        GROUP BY COALESCE(d.name, 'Dept ' || ti.merchandise_code)
    """, (target_date,))
    dept_sales = dict(cur.fetchall())

    for dept_name, row_idx in dept_mapping.items():
        amt = 0.0
        for d_key, d_val in dept_sales.items():
            if dept_name.lower() in d_key.lower():
                amt += d_val
        ws[f'B{row_idx}'] = round(amt, 2)

    # Lottery & Other
    cur.execute("""
        SELECT ti.description, SUM(ti.total_amount)
        FROM transaction_items ti
        JOIN transactions t ON ti.transaction_id = t.id
        WHERE t.business_date = ? AND (ti.description LIKE '%lotto%' OR ti.description LIKE '%scratch%' OR ti.description LIKE '%gas card%')
        GROUP BY ti.description
    """, (target_date,))
    lottery_rows = cur.fetchall()

    scratch_amt = 0.0
    lotto_amt = 0.0
    gas_card_amt = 0.0

    for desc, amt in lottery_rows:
        d_lower = desc.lower()
        if 'scratch' in d_lower: scratch_amt += amt
        elif 'lotto' in d_lower: lotto_amt += amt
        elif 'gas card' in d_lower: gas_card_amt += amt

    ws['B26'] = round(gas_card_amt, 2)
    ws['B27'] = round(scratch_amt, 2)
    ws['B28'] = round(lotto_amt, 2)

    # Sales Tax
    cur.execute("SELECT SUM(tax_amount) FROM transactions WHERE business_date = ?", (target_date,))
    tax_row = cur.fetchone()
    sales_tax = tax_row[0] if tax_row and tax_row[0] else 0.0
    ws['B29'] = round(sales_tax, 2)

    # 4. Official FGM Fuel Sales
    fgm_official = get_official_fgm_fuel(staging_dir, target_date)
    if fgm_official:
        print(f"Using Official Passport FGM Storewide Net Fuel Data for {target_date}")
        reg_gal = fgm_official['001']['gal']
        reg_amt = fgm_official['001']['amt'] + fgm_official['001']['disc']

        plus_gal = fgm_official['002']['gal']
        plus_amt = fgm_official['002']['amt'] + fgm_official['002']['disc']

        super_gal = fgm_official['003']['gal']
        super_amt = fgm_official['003']['amt'] + fgm_official['003']['disc']
    else:
        cur.execute("""
            SELECT ti.description, SUM(ti.total_amount), SUM(ti.quantity)
            FROM transaction_items ti
            JOIN transactions t ON ti.transaction_id = t.id
            WHERE t.business_date = ? AND ti.item_type = 'fuel'
            GROUP BY ti.description
        """, (target_date,))
        fuel_rows = cur.fetchall()

        reg_amt, reg_gal = 0.0, 0.0
        plus_amt, plus_gal = 0.0, 0.0
        super_amt, super_gal = 0.0, 0.0

        for desc, amt, gal in fuel_rows:
            d_upper = desc.upper()
            if 'REGULAR' in d_upper or 'UNLEADED' in d_upper:
                reg_amt += amt; reg_gal += gal
            elif 'PLUS' in d_upper or 'MIDGRADE' in d_upper:
                plus_amt += amt; plus_gal += gal
            elif 'PREMIUM' in d_upper or 'SUPER' in d_upper or 'DIESEL' in d_upper or 'FUEL' in d_upper:
                super_amt += amt; super_gal += gal

    ws['E5'] = round(reg_amt, 2); ws['F5'] = round(reg_gal, 3)
    ws['E6'] = round(plus_amt, 2); ws['F6'] = round(plus_gal, 3)
    ws['E7'] = round(super_amt, 2); ws['F7'] = round(super_gal, 3)
    ws['E8'] = round(reg_amt + plus_amt + super_amt, 2)

    # Payments & Deposit
    cur.execute("""
        SELECT p.tender_code, SUM(p.amount)
        FROM payments p
        JOIN transactions t ON p.transaction_id = t.id
        WHERE t.business_date = ?
        GROUP BY p.tender_code
    """, (target_date,))
    pmt_rows = dict(cur.fetchall())

    cash_amt = pmt_rows.get('cash', 0.0)
    cc_amt = sum(v for k, v in pmt_rows.items() if k != 'cash')

    ws['I5'] = round(cash_amt, 2)
    ws['I6'] = 0.0
    ws['I10'] = round(cc_amt, 2)

    wb.save(output_path)
    print(f"Generated daily recon spreadsheet: {output_path}")

    # Update Master Log
    master_log_path = os.path.join(out_dir, 'DailyReconciliation_Master_Log.xlsx')
    update_master_log(master_log_path, target_date, ws)

    conn.close()
    return output_path

def update_master_log(log_path, date_str, ws_source):
    if os.path.exists(log_path):
        wb_master = openpyxl.load_workbook(log_path)
        ws_master = wb_master.active
    else:
        wb_master = openpyxl.Workbook()
        ws_master = wb_master.active
        ws_master.title = "Reconciliation Master Log"
        headers = ["Date", "Auto Parts", "Beer", "Candy", "Cig Cartons", "Cigs", "Coffee", "Deli", "Edible", "Fountain", "GROC No Tax", "Ice", "Liquor", "Non-Edible", "Snacks", "Soda", "Tobacco", "Vapes", "HBA", "Gas Card", "Scratch", "Machine Lotto", "Sales Tax", "Regular ($)", "Regular (Gal)", "Plus ($)", "Plus (Gal)", "Super ($)", "Super (Gal)", "Gross Fuel ($)", "Total Inside ($)", "Total Sales ($)", "Cash ($)", "Credit Cards ($)"]
        ws_master.append(headers)

    existing_row = None
    for r in range(2, ws_master.max_row + 1):
        if str(ws_master.cell(r, 1).value) == str(date_str):
            existing_row = r
            break

    row_data = [
        date_str,
        ws_source['B5'].value, ws_source['B6'].value, ws_source['B7'].value, ws_source['B8'].value,
        ws_source['B9'].value, ws_source['B10'].value, ws_source['B11'].value, ws_source['B12'].value,
        ws_source['B13'].value, ws_source['B14'].value, ws_source['B15'].value, ws_source['B16'].value,
        ws_source['B17'].value, ws_source['B18'].value, ws_source['B19'].value, ws_source['B20'].value,
        ws_source['B21'].value, ws_source['B22'].value, ws_source['B26'].value, ws_source['B27'].value,
        ws_source['B28'].value, ws_source['B29'].value, ws_source['E5'].value, ws_source['F5'].value,
        ws_source['E6'].value, ws_source['F6'].value, ws_source['E7'].value, ws_source['F7'].value,
        ws_source['E8'].value,
        round(sum(ws_source[f'B{r}'].value or 0 for r in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,26,27,28,29]), 2),
        round((ws_source['E8'].value or 0) + sum(ws_source[f'B{r}'].value or 0 for r in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,26,27,28,29]), 2),
        ws_source['I5'].value, ws_source['I10'].value
    ]

    if existing_row:
        for c, val in enumerate(row_data, 1):
            ws_master.cell(existing_row, c, val)
    else:
        ws_master.append(row_data)

    wb_master.save(log_path)
    print(f"Master log updated: {log_path}")

if __name__ == '__main__':
    target = sys.argv[1] if len(sys.argv) > 1 else None
    generate_recon(target)
