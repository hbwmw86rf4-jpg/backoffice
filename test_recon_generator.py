import openpyxl
import sqlite3
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

db_path = r'C:\Users\shell\Documents\office\backoffice\data\backoffice.db'
template_path = r'C:\Users\sandh\Downloads\RochesterShell_DailyRecon.xlsx'
output_path = r'C:\Users\sandh\Downloads\RochesterShell_DailyRecon_2026-08-15.xlsx'

if not os.path.exists(template_path):
    print(f"Template not found: {template_path}")
    exit(1)

conn = sqlite3.connect(db_path)
cur = conn.cursor()

target_date = '2026-08-15'
print(f"Generating Daily Reconciliation Spreadsheet for date: {target_date}...")

wb = openpyxl.load_workbook(template_path)
ws = wb['Daily Reconciliation']

# 1. Date
ws['I2'] = target_date

# 2. Inside Sales by Department
dept_mapping = {
    'Auto Parts': 5,
    'Beer': 6,
    'Candy': 7,
    'Cig Cartons': 8,
    'Cigs': 9,
    'Coffee': 10,
    'Deli': 11,
    'Edible': 12,
    'Fountain': 13,
    'GROC No Tax': 14,
    'Ice': 15,
    'Liquor': 16,
    'Non-Edible': 17,
    'Snacks': 18,
    'Soda': 19,
    'Tobacco': 20,
    'Vapes': 21,
    'HBA': 22
}

# Fetch department sales
cur.execute("""
    SELECT COALESCE(d.name, 'Dept ' || ti.merchandise_code), SUM(ti.total_amount)
    FROM transaction_items ti
    JOIN transactions t ON ti.transaction_id = t.id
    LEFT JOIN pricebook pb ON ti.upc = pb.upc
    LEFT JOIN departments d ON d.id = COALESCE(pb.department_id, CAST(ti.merchandise_code AS INTEGER))
    WHERE t.business_date = ? AND ti.item_type = 'cstore'
    GROUP BY COALESCE(d.name, 'Dept ' || ti.merchandise_code)
""", (target_date,))
dept_sales = dict(cur.fetchall())

print("\nDepartment Sales:")
for dept_name, row_idx in dept_mapping.items():
    amt = 0.0
    for d_key, d_val in dept_sales.items():
        if dept_name.lower() in d_key.lower():
            amt += d_val
    ws[f'B{row_idx}'] = amt
    print(f"  {dept_name:15} (B{row_idx}): ${amt:.2f}")

# 3. Lottery & Other
cur.execute("""
    SELECT ti.description, SUM(ti.total_amount)
    FROM transaction_items ti
    JOIN transactions t ON ti.transaction_id = t.id
    WHERE t.business_date = ? AND (ti.description LIKE '%lotto%' OR ti.description LIKE '%scratch%' OR ti.description LIKE '%gas card%')
    GROUP BY ti.description
""", (target_date,))
lottery_rows = cur.fetchall()

scratch_amt = 0.0
lotto_amt = 0.0
gas_card_amt = 0.0

for desc, amt in lottery_rows:
    if 'scratch' in desc.lower(): scratch_amt += amt
    elif 'lotto' in desc.lower(): lotto_amt += amt
    elif 'gas card' in desc.lower(): gas_card_amt += amt

ws['B26'] = gas_card_amt
ws['B27'] = scratch_amt
ws['B28'] = lotto_amt

# Sales Tax
cur.execute("SELECT SUM(tax_amount) FROM transactions WHERE business_date = ?", (target_date,))
tax_row = cur.fetchone()
sales_tax = tax_row[0] if tax_row and tax_row[0] else 0.0
ws['B29'] = sales_tax

# 4. Fuel Sales
cur.execute("""
    SELECT ti.description, SUM(ti.total_amount), SUM(ti.quantity)
    FROM transaction_items ti
    JOIN transactions t ON ti.transaction_id = t.id
    WHERE t.business_date = ? AND ti.item_type = 'fuel'
    GROUP BY ti.description
""", (target_date,))
fuel_rows = cur.fetchall()

reg_amt, reg_gal = 0.0, 0.0
plus_amt, plus_gal = 0.0, 0.0
super_amt, super_gal = 0.0, 0.0

for desc, amt, gal in fuel_rows:
    d_upper = desc.upper()
    if 'REGULAR' in d_upper or 'UNLEADED' in d_upper:
        reg_amt += amt; reg_gal += gal
    elif 'PLUS' in d_upper or 'MIDGRADE' in d_upper:
        plus_amt += amt; plus_gal += gal
    elif 'PREMIUM' in d_upper or 'SUPER' in d_upper or 'DIESEL' in d_upper or 'FUEL' in d_upper:
        super_amt += amt; super_gal += gal

ws['E5'] = reg_amt; ws['F5'] = reg_gal
ws['E6'] = plus_amt; ws['F6'] = plus_gal
ws['E7'] = super_amt; ws['F7'] = super_gal

# Fuel net / 1 Dept
ws['E8'] = reg_amt + plus_amt + super_amt

# 5. Payments / Deposit / CC
cur.execute("""
    SELECT p.tender_code, SUM(p.amount)
    FROM payments p
    JOIN transactions t ON p.transaction_id = t.id
    WHERE t.business_date = ?
    GROUP BY p.tender_code
""", (target_date,))
pmt_rows = dict(cur.fetchall())

cash_amt = pmt_rows.get('cash', 0.0)
cc_amt = sum(v for k, v in pmt_rows.items() if k != 'cash')

ws['I5'] = cash_amt
ws['I6'] = 0.0  # Checks
ws['I10'] = cc_amt

wb.save(output_path)
print(f"\nSuccessfully generated reconciliation spreadsheet at: {output_path}")

conn.close()
